"""
主排程迴圈（從 backend/ 目錄執行：python scheduler.py）
- 每 60 分鐘執行一次掃描
- 在 4H K 線收盤時（UTC 00/04/08/12/16/20）立即額外觸發
- 同一幣種 4 小時內不重複推播
"""

import csv
import json
import time
from datetime import datetime, timezone, timedelta
from pathlib import Path

from app.services.data_ingestion.binance import get_contracts, get_klines
from app.services.strategies.scanner     import scan_symbol
from app.services.strategies.indicators  import ema_snapshot
from app.services.scoring.scorer         import score_setup, passes_threshold
from app.services.paper_trader           import PaperTrader
from app.core.config import (
    DATA_DIR, STATE_FILE, SIGNALS_LOG, SIGNALS_JSONL,
    EQUITY_JSONL, TRADE_RECORDS_DIR, CLOSED_TRADES_JSONL, PAPER_FILE,
)

TW_TZ           = timezone(timedelta(hours=8))
TRADE_CSV       = DATA_DIR / "trade_history.csv"
SCAN_INTERVAL   = 60 * 60
DEDUP_WINDOW    = 4 * 60 * 60
KLINES_4H_LIMIT = 250
KLINES_1H_LIMIT = 100

BOT_START_TIME: float = 0.0

TRADE_CSV_FIELDS = [
    "symbol", "direction", "score", "entry", "exit",
    "stop_loss", "target1", "target2", "contracts", "pnl",
    "reason", "full_close", "open_time", "close_time",
]


def _fmt_tw(ts: float) -> str:
    return datetime.fromtimestamp(ts, tz=TW_TZ).strftime("%Y/%m/%d %H:%M:%S")


def _duration_str(start: float, stop: float) -> str:
    secs = int(stop - start)
    h, rem = divmod(secs, 3600)
    m, s   = divmod(rem, 60)
    if h > 0:
        return f"{h}h {m}m"
    if m > 0:
        return f"{m}m {s}s"
    return f"{s}s"


def archive_session(start_ts: float, stop_ts: float, trader: PaperTrader) -> None:
    """關機時將本次交易紀錄匯出為 Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    TRADE_RECORDS_DIR.mkdir(parents=True, exist_ok=True)
    file_name = datetime.fromtimestamp(start_ts, tz=TW_TZ).strftime("%Y-%m-%d_%H-%M") + ".xlsx"
    dest      = TRADE_RECORDS_DIR / file_name

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "交易明細"
    col_map = [
        ("開倉時間", "open_ms"), ("平倉時間", "close_ms"), ("交易對", "symbol"),
        ("方向", "direction"), ("策略", "strategy"), ("評分", "score"),
        ("進場價", "entry"), ("出場價", "exit"), ("止損", "stop_loss"),
        ("TP1", "target1"), ("TP2", "target2"), ("張數", "contracts"),
        ("損益 $", "pnl"), ("出場原因", "reason"), ("完整平倉", "full_close"),
    ]

    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col_idx, (label, _) in enumerate(col_map, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, t in enumerate(trader.trade_history, 2):
        pnl      = t.get("pnl", 0)
        row_fill = PatternFill("solid", fgColor="C6EFCE" if pnl >= 0 else "FFC7CE")
        for col_idx, (_, key) in enumerate(col_map, 1):
            val = t.get(key)
            if key in ("open_ms", "close_ms") and isinstance(val, (int, float)):
                val = datetime.fromtimestamp(val / 1000, tz=TW_TZ).strftime("%Y-%m-%d %H:%M:%S")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            if key == "pnl":
                cell.number_format = '+#,##0.00;-#,##0.00'

    for col_idx in range(1, len(col_map) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].auto_size = True
    ws.freeze_panes = "A2"

    stats = trader.get_stats()
    ws2   = wb.create_sheet("摘要")
    summary_rows = [
        ("開機時間",   _fmt_tw(start_ts)),
        ("關機時間",   _fmt_tw(stop_ts)),
        ("執行時長",   _duration_str(start_ts, stop_ts)),
        ("初始資金",   f"${trader.initial_balance:,.2f}"),
        ("最終餘額",   f"${stats.get('current_balance', 0):,.2f}"),
        ("總損益",     f"${stats.get('total_pnl', 0):+,.2f}"),
        ("報酬率",     f"{stats.get('total_return_pct', 0):+.2f}%"),
        ("最大回撤",   f"{stats.get('max_drawdown_pct', 0):.2f}%"),
        ("勝率",       f"{stats.get('win_rate_pct', 0):.1f}%"),
        ("已平倉",     stats.get("full_closes", 0)),
        ("勝",         stats.get("wins", 0)),
        ("敗",         stats.get("losses", 0)),
        ("獲利因子",   stats.get("profit_factor")),
    ]
    key_font = Font(bold=True)
    key_fill = PatternFill("solid", fgColor="DDEEFF")
    for r, (label, val) in enumerate(summary_rows, 1):
        k = ws2.cell(row=r, column=1, value=label)
        k.font = key_font
        k.fill = key_fill
        ws2.cell(row=r, column=2, value=val)
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 22

    wb.save(dest)
    print(f"[ARCHIVE] 交易紀錄已存至 data/trade_records/{file_name}")


# ── 去重狀態管理 ──────────────────────────────────────────────

def load_state() -> dict:
    if STATE_FILE.exists():
        try:
            with open(STATE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def save_state(state: dict) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)


def is_duplicate(state: dict, symbol: str) -> bool:
    last_ts = state.get(symbol)
    if last_ts is None:
        return False
    return (time.time() - last_ts) < DEDUP_WINDOW


def mark_alerted(state: dict, symbol: str) -> None:
    state[symbol] = time.time()


def prune_state(state: dict) -> dict:
    now = time.time()
    return {k: v for k, v in state.items() if now - v < DEDUP_WINDOW}


def log_signal(result: dict, score: float) -> None:
    entry = {
        "symbol":    result["symbol"],
        "direction": result["direction"],
        "strategy":  result.get("strategy", "EMA_CONVERGENCE"),
        "score":     score,
        "entry":     result["levels"]["entry"],
        "stop_loss": result["levels"]["stop_loss"],
        "target1":   result["levels"]["target1"],
        "target2":   result["levels"]["target2"],
        "bandwidth": result["convergence"]["bandwidth"],
        "time":      datetime.now(TW_TZ).strftime("%Y/%m/%d %H:%M"),
    }
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        existing: list = json.loads(SIGNALS_LOG.read_text(encoding="utf-8")) if SIGNALS_LOG.exists() else []
        existing.append(entry)
        SIGNALS_LOG.write_text(
            json.dumps(existing[-300:], ensure_ascii=False, indent=2), encoding="utf-8"
        )
    except Exception:
        pass
    try:
        with open(SIGNALS_JSONL, "a", encoding="utf-8") as f:
            f.write(json.dumps(entry, ensure_ascii=False) + "\n")
    except Exception:
        pass


def export_trades_csv(trader: PaperTrader) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with open(TRADE_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=TRADE_CSV_FIELDS, extrasaction="ignore")
        writer.writeheader()
        for t in trader.trade_history:
            row = dict(t)
            row["open_time"]  = datetime.fromtimestamp(t["open_ms"]  / 1000, tz=TW_TZ).strftime("%Y-%m-%d %H:%M:%S")
            row["close_time"] = datetime.fromtimestamp(t["close_ms"] / 1000, tz=TW_TZ).strftime("%Y-%m-%d %H:%M:%S")
            writer.writerow(row)


def log_closed_trades(events: list[dict]) -> None:
    if not events:
        return
    TRADE_RECORDS_DIR.mkdir(parents=True, exist_ok=True)
    with open(CLOSED_TRADES_JSONL, "a", encoding="utf-8") as f:
        for ev in events:
            rec = dict(ev)
            if isinstance(rec.get("open_ms"), (int, float)):
                rec["open_time"]  = datetime.fromtimestamp(rec["open_ms"]  / 1000, tz=TW_TZ).strftime("%Y-%m-%d %H:%M:%S")
            if isinstance(rec.get("close_ms"), (int, float)):
                rec["close_time"] = datetime.fromtimestamp(rec["close_ms"] / 1000, tz=TW_TZ).strftime("%Y-%m-%d %H:%M:%S")
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")


def append_equity_snapshot(trader: PaperTrader) -> None:
    stats = trader.get_stats()
    rec   = {
        "time":           datetime.now(TW_TZ).strftime("%Y-%m-%d %H:%M:%S"),
        "balance":        stats["current_balance"],
        "total_pnl":      stats.get("total_pnl", 0.0),
        "open_positions": stats.get("open_positions", len(trader.positions)),
    }
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with open(EQUITY_JSONL, "a", encoding="utf-8") as f:
        f.write(json.dumps(rec, ensure_ascii=False) + "\n")


# ── 4H 收盤時間偵測 ───────────────────────────────────────────

def next_4h_close_utc() -> datetime:
    now         = datetime.now(timezone.utc)
    hour_bucket = (now.hour // 4) * 4
    base        = now.replace(hour=hour_bucket, minute=0, second=0, microsecond=0)
    candidate   = base + timedelta(hours=4)
    if candidate <= now:
        candidate += timedelta(hours=4)
    return candidate


def _print_open_detail(result: dict, score: float) -> None:
    strat_map = {
        "EMA_CONVERGENCE":    "EMA收斂突破",
        "EMA_PULLBACK":       "EMA30回測",
        "STRUCTURE_BREAKOUT": "結構突破回測",
    }
    strategy   = result.get("strategy", "?")
    strat_name = strat_map.get(strategy, strategy)
    lvl        = result["levels"]
    conv       = result["convergence"]
    vol        = result["vol_ratio"]
    conf       = result["confirm_1h"]
    direction  = result["direction"]

    print(f"\n  ┌─ [開倉] {result['symbol']:22s} {direction:5s}  score={score}  策略={strat_name}")
    print(f"  │  進場={lvl['entry']:.6g}  SL={lvl['stop_loss']:.6g}  TP1={lvl['target1']:.6g}  TP2={lvl['target2']:.6g}")
    if strategy == "EMA_CONVERGENCE":
        print(f"  │  帶寬={conv['bandwidth']:.2f}% 壓縮{conv['compression_bars']}根  量比={vol:.1f}×  實體={conf['body_ratio']*100:.0f}%")
    elif strategy == "EMA_PULLBACK":
        print(f"  │  回測EMA30後反彈確認  量比={vol:.1f}×  實體={conf['body_ratio']*100:.0f}%")
    elif strategy == "STRUCTURE_BREAKOUT":
        print(f"  │  結構突破回測確認  量比={vol:.1f}×  實體={conf['body_ratio']*100:.0f}%")
    print(f"  └{'─'*60}")


# ── 核心掃描函式 ──────────────────────────────────────────────

def run_scan() -> None:
    now_tw = datetime.now(TW_TZ).strftime("%Y/%m/%d %H:%M TWN")
    print(f"\n[掃描開始] {now_tw}")

    symbols = get_contracts()
    if not symbols:
        print("[錯誤] 無法取得合約清單，本次掃描跳過")
        return

    print(f"[資訊] 共取得 {len(symbols)} 個交易對")

    # Regime Filter：BTC 4H EMA15 vs EMA60
    btc_regime_bull = True
    btc_candles_4h  = get_klines("BTCUSDT", "4h", KLINES_4H_LIMIT)
    if btc_candles_4h and len(btc_candles_4h) >= 60:
        btc_emas = ema_snapshot(btc_candles_4h)
        if btc_emas:
            _idx = len(btc_candles_4h) - 1
            _e15 = btc_emas["ema15"][_idx]
            _e60 = btc_emas["ema60"][_idx]
            if _e15 is not None and _e60 is not None:
                btc_regime_bull = _e15 > _e60
    regime_str = "多頭" if btc_regime_bull else "空頭（山寨多單封鎖）"
    print(f"[Regime] BTC 4H EMA15 {'>' if btc_regime_bull else '<'} EMA60 → {regime_str}")

    state  = prune_state(load_state())
    trader = PaperTrader.load()

    qualified:         list[tuple]       = []
    converging:        int               = 0
    total:             int               = len(symbols)
    latest_bar_4h:     dict[str, dict]   = {}

    for i, symbol in enumerate(symbols, 1):
        candles_4h = get_klines(symbol, "4h", KLINES_4H_LIMIT)
        candles_1h = get_klines(symbol, "1h", KLINES_1H_LIMIT)
        if candles_4h is None or candles_1h is None:
            continue

        latest_bar_4h[symbol] = candles_4h[-1]

        result = scan_symbol(symbol, candles_4h, candles_1h, btc_regime_bull)
        if result is None:
            continue

        converging += 1
        score = score_setup(result, result.get("candle_time_ms"))
        if not passes_threshold(score):
            continue

        if is_duplicate(state, symbol):
            print(f"  [跳過重複] {symbol}  score={score}")
            continue

        strat = result.get("strategy", "EMA_CONVERGENCE")
        print(f"  [訊號] {symbol}  {result['direction']}  [{strat}]  score={score}")
        if len(trader.positions) >= trader.max_positions:
            print(f"  [跳過] 已達持倉上限 {trader.max_positions}，{symbol} 排隊等待")

        qualified.append((result, score))
        mark_alerted(state, symbol)
        log_signal(result, score)

        if i % 20 == 0:
            print(f"  ... 已掃描 {i}/{total}")

    save_state(state)

    # 對持倉中但本輪未掃描的幣種補取最新 K 線
    for sym in list(trader.positions.keys()):
        if sym not in latest_bar_4h:
            extra = get_klines(sym, "4h", 5)
            if extra:
                latest_bar_4h[sym] = extra[-1]

    exit_events = trader.update_positions(latest_bar_4h)
    log_closed_trades(exit_events)
    for ev in exit_events:
        print(f"  [紙倉出場] {ev['symbol']:20s}  {ev['reason']}  PnL: ${ev['pnl']:>+.2f}")

    qualified.sort(key=lambda x: x[1], reverse=True)
    for result, score in qualified:
        if trader.open_position(result, score):
            _print_open_detail(result, score)

    trader.save()
    export_trades_csv(trader)
    append_equity_snapshot(trader)
    trader.print_report()

    if qualified:
        print(f"[完成] {len(qualified)} 個達標訊號")
    else:
        print(f"[完成] 無達標訊號（收斂中：{converging}）")


# ── 排程主迴圈 ────────────────────────────────────────────────

def main() -> None:
    global BOT_START_TIME
    BOT_START_TIME = time.time()

    print("=" * 55)
    print("  Crypto Quant Signal Platform — Scheduler 啟動")
    print("  交易所：Binance Futures")
    print("  掃描間隔：60 分鐘")
    print("  額外觸發：4H K 線收盤時")
    print("=" * 55)

    try:
        run_scan()
        last_scan_ts   = time.time()
        last_4h_window = datetime.now(timezone.utc).hour // 4

        while True:
            time.sleep(30)
            now_utc       = datetime.now(timezone.utc)
            current_4h_w  = now_utc.hour // 4
            elapsed       = time.time() - last_scan_ts

            triggered_by_interval = elapsed >= SCAN_INTERVAL
            triggered_by_4h_close = current_4h_w != last_4h_window

            if triggered_by_interval or triggered_by_4h_close:
                if triggered_by_4h_close:
                    print(f"[觸發] 新 4H K 線收盤（UTC 窗口 {current_4h_w * 4:02d}:00）")
                last_4h_window = current_4h_w
                last_scan_ts   = time.time()
                run_scan()

    except KeyboardInterrupt:
        print("\nStopped by user.")
    finally:
        stop_ts = time.time()
        try:
            trader = PaperTrader.load()
            archive_session(BOT_START_TIME, stop_ts, trader)
        except Exception as e:
            print(f"[ARCHIVE] 匯出失敗：{e}")


if __name__ == "__main__":
    main()
