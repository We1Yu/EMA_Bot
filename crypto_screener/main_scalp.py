"""
High-Frequency Virtual Trading Bot — BingX USDT-M Perpetuals
=============================================================

Two independent loops:
  - Position loop (every SCALP_CHECK_INTERVAL_SECS, default 12s):
    fetches the current price for every open paper position and checks
    stop-loss / take-profit immediately, so exits trigger fast.
  - Scan loop (aligned to 5-minute candle close, +5s buffer):
    re-runs the full crypto_screener scoring pipeline on SCALP_INTERVAL
    candles (default 5m) across all BingX perpetuals and opens new
    paper positions for fresh signals.

State is written to scalp_state.json and paper_account_scalp.json so
web_app.py can render a live dashboard.

Usage:
  python main_scalp.py
"""

import asyncio
import csv
import json
import logging
import shutil
import time
from datetime import datetime, timezone, timedelta
from pathlib import Path

import aiohttp

import pandas as _pd
import bingx_source as source
import bingx_source as _bingx
from scanner import run_scan
from paper_account import PaperAccount
from discord_alert import send_daily_report
from indicators import calc_sma as _calc_sma, calc_adx as _calc_adx

from config import (
    SCALP_INTERVAL, SCALP_CHECK_INTERVAL_SECS,
    SCALP_COOLDOWN_SECS, SCALP_PAPER_INITIAL_BALANCE, SCALP_PAPER_RISK_PCT,
    SCALP_DISCORD_WEBHOOK,
    SCALP_MIN_SCORE, SCALP_MIN_RR, SCALP_MAX_POSITIONS,
    SCALP_MAX_ENTRY_DRIFT, SCALP_SYMBOL_BLOCKLIST,
    BTC_SYMBOL, BTC_MA_PERIOD, BTC_ADX_PAUSE,
    FUNDING_EXTREME,
)

TW_TZ      = timezone(timedelta(hours=8))
PAPER_FILE = Path(__file__).parent / "paper_account_scalp.json"
STATE_FILE = Path(__file__).parent / "scalp_state.json"

BOT_START_TIME: float = 0.0

TRADE_CSV     = Path(__file__).parent / "trade_history_scalp.csv"
SIGNALS_JSONL = Path(__file__).parent / "signals_history_scalp.jsonl"
EQUITY_JSONL  = Path(__file__).parent / "equity_history_scalp.jsonl"
TRADES_JSONL  = Path(__file__).parent / "trades_scalp.jsonl"

SESSIONS_JSONL = Path(__file__).parent / "sessions_log.jsonl"
SESSIONS_DIR   = Path(__file__).parent / "sessions"
BOT_STOP_TIME: float = 0.0

TRADE_CSV_FIELDS = [
    "symbol", "strategy", "direction", "tier", "score", "entry", "exit",
    "stop_loss", "target_1", "target_2", "contracts", "pnl",
    "reason", "full_close", "open_time", "close_time",
]

log = logging.getLogger(__name__)


def _fmt_tw(ts: float) -> str:
    return datetime.fromtimestamp(ts, tz=TW_TZ).strftime("%Y/%m/%d %H:%M:%S")


def _duration_str(start: float, stop: float) -> str:
    secs = int(stop - start)
    h, rem = divmod(secs, 3600)
    m, s   = divmod(rem, 60)
    if h > 0:
        return f"{h}h {m}m"
    if m > 0:
        return f"{m}m {s}s"
    return f"{s}s"


def append_session_log(event: str, ts: float, extra: dict | None = None) -> None:
    rec = {"event": event, "time": _fmt_tw(ts), "ts": ts, **(extra or {})}
    with open(SESSIONS_JSONL, "a", encoding="utf-8") as f:
        f.write(json.dumps(rec, ensure_ascii=False) + "\n")


def archive_session(start_ts: float, stop_ts: float, stats: dict) -> None:
    """關機時將本次交易紀錄匯出為 Excel，存至 sessions/<start_time>.xlsx。"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter

    SESSIONS_DIR.mkdir(parents=True, exist_ok=True)
    file_name = datetime.fromtimestamp(start_ts, tz=TW_TZ).strftime("%Y-%m-%d_%H-%M") + ".xlsx"
    dest = SESSIONS_DIR / file_name

    wb = openpyxl.Workbook()

    # ── 分頁 1：交易明細 ────────────────────────────────────────
    ws = wb.active
    ws.title = "交易明細"

    trades: list[dict] = []
    if TRADES_JSONL.exists():
        for line in TRADES_JSONL.read_text(encoding="utf-8").splitlines():
            if line.strip():
                trades.append(json.loads(line))

    col_map = [
        ("開倉時間",   "open_time"),
        ("平倉時間",   "close_time"),
        ("交易對",     "symbol"),
        ("策略",       "strategy"),
        ("方向",       "direction"),
        ("Tier",       "tier"),
        ("評分",       "score"),
        ("進場價",     "entry"),
        ("出場價",     "exit"),
        ("止損",       "stop_loss"),
        ("TP1",        "target_1"),
        ("TP2",        "target_2"),
        ("張數",       "contracts"),
        ("損益 $",     "pnl"),
        ("出場原因",   "reason"),
        ("完整平倉",   "full_close"),
    ]

    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(bold=True, color="FFFFFF")

    for col_idx, (label, _) in enumerate(col_map, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, t in enumerate(trades, 2):
        pnl = t.get("pnl", 0)
        row_fill = PatternFill("solid", fgColor="C6EFCE" if pnl >= 0 else "FFC7CE")
        for col_idx, (_, key) in enumerate(col_map, 1):
            val = t.get(key)
            if key in ("open_time", "close_time") and isinstance(val, (int, float)):
                val = datetime.fromtimestamp(val, tz=TW_TZ).strftime("%Y-%m-%d %H:%M:%S")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            if key == "pnl":
                cell.number_format = '+#,##0.00;-#,##0.00'

    for col_idx in range(1, len(col_map) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].auto_size = True

    ws.freeze_panes = "A2"

    # ── 分頁 2：摘要 ────────────────────────────────────────────
    ws2 = wb.create_sheet("摘要")
    summary_rows = [
        ("開機時間",   _fmt_tw(start_ts)),
        ("關機時間",   _fmt_tw(stop_ts)),
        ("執行時長",   _duration_str(start_ts, stop_ts)),
        ("初始資金",   f"${SCALP_PAPER_INITIAL_BALANCE:,.2f}"),
        ("最終餘額",   f"${stats.get('current_balance', 0):,.2f}"),
        ("總損益",     f"${stats.get('total_pnl', 0):+,.2f}"),
        ("報酬率",     f"{stats.get('return_pct', 0):+.2f}%"),
        ("最大回撤",   f"{stats.get('max_drawdown_pct', 0):.2f}%"),
        ("勝率",       f"{stats.get('win_rate', 0):.1f}%"),
        ("已平倉",     stats.get("full_closes", 0)),
        ("勝",         stats.get("wins", 0)),
        ("敗",         stats.get("losses", 0)),
        ("獲利因子",   stats.get("profit_factor")),
    ]
    key_font  = Font(bold=True)
    key_fill  = PatternFill("solid", fgColor="DDEEFF")
    for r, (label, val) in enumerate(summary_rows, 1):
        k = ws2.cell(row=r, column=1, value=label)
        k.font = key_font
        k.fill = key_fill
        ws2.cell(row=r, column=2, value=val)
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 22

    wb.save(dest)
    print(f"[ARCHIVE] 交易紀錄已存至 sessions/{file_name}")


def save_state(account: PaperAccount, signals: list[dict], total_scanned: int) -> None:
    state = {
        "updated":       time.time(),
        "start_time":    BOT_START_TIME,
        "stop_time":     BOT_STOP_TIME if BOT_STOP_TIME else None,
        "interval":      SCALP_INTERVAL,
        "total_scanned": total_scanned,
        "stats":         account.get_stats(),
        "signals":       signals[:20],
    }
    STATE_FILE.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")


def export_trades_csv(account: PaperAccount) -> None:
    """Rewrite the full closed-trade history as CSV for analysis."""
    with open(TRADE_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=TRADE_CSV_FIELDS, extrasaction="ignore")
        writer.writeheader()
        for t in account.history:
            row = dict(t)
            row["open_time"]  = datetime.fromtimestamp(t["open_time"],  tz=TW_TZ).strftime("%Y-%m-%d %H:%M:%S")
            row["close_time"] = datetime.fromtimestamp(t["close_time"], tz=TW_TZ).strftime("%Y-%m-%d %H:%M:%S")
            writer.writerow(row)


def append_signals_jsonl(signals: list[dict], total_scanned: int) -> None:
    """Append every signal from this scan (not just the top 20) for later analysis."""
    ts = datetime.now(TW_TZ).strftime("%Y-%m-%d %H:%M:%S")
    with open(SIGNALS_JSONL, "a", encoding="utf-8") as f:
        for sg in signals:
            rec = {"time": ts, "total_scanned": total_scanned, **sg}
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")


def append_trades_jsonl(events: list[dict], account: PaperAccount) -> None:
    """Append each exit event to an immutable JSONL log for permanent record."""
    ts = datetime.now(TW_TZ).strftime("%Y-%m-%d %H:%M:%S")
    stats = account.get_stats()
    with open(TRADES_JSONL, "a", encoding="utf-8") as f:
        for ev in events:
            rec = {
                "time":    ts,
                "balance": stats["current_balance"],
                **ev,
            }
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")


def append_equity_snapshot(account: PaperAccount) -> None:
    """Append a balance snapshot so an equity curve can be plotted later."""
    ts    = datetime.now(TW_TZ).strftime("%Y-%m-%d %H:%M:%S")
    stats = account.get_stats()
    rec = {
        "time":            ts,
        "balance":         stats["current_balance"],
        "total_pnl":       stats["total_pnl"],
        "open_positions":  stats["open_positions"],
    }
    with open(EQUITY_JSONL, "a", encoding="utf-8") as f:
        f.write(json.dumps(rec, ensure_ascii=False) + "\n")


# ── BTC trend helper ─────────────────────────────────────────

async def fetch_btc_trend(session: aiohttp.ClientSession) -> tuple[str, bool]:
    """
    Returns (btc_bias, extreme_volatility).
    btc_bias: "BULLISH" | "BEARISH" | "UNKNOWN"
    extreme_volatility: True if BTC ADX > BTC_ADX_PAUSE (all entries paused)
    """
    try:
        raw = await _bingx.fetch_klines(session, BTC_SYMBOL, "5m", BTC_MA_PERIOD + 20)
        if not raw or len(raw) < BTC_MA_PERIOD + 5:
            return "UNKNOWN", False
        df    = _pd.DataFrame(raw)
        close = _pd.to_numeric(df["close"], errors="coerce")
        high  = _pd.to_numeric(df["high"],  errors="coerce")
        low   = _pd.to_numeric(df["low"],   errors="coerce")
        ma60  = _calc_sma(close, BTC_MA_PERIOD).iloc[-1]
        adx_s, _, _ = _calc_adx(high, low, close, 14)
        adx   = adx_s.iloc[-1]
        if _pd.isna(ma60) or _pd.isna(adx):
            return "UNKNOWN", False
        extreme  = float(adx) > BTC_ADX_PAUSE
        btc_bias = "BULLISH" if float(close.iloc[-1]) > float(ma60) else "BEARISH"
        return btc_bias, extreme
    except Exception as e:
        log.debug("fetch_btc_trend error: %s", e)
        return "UNKNOWN", False


# ── Loop 1: fast position checks ─────────────────────────────

async def position_loop(account: PaperAccount, session: aiohttp.ClientSession) -> None:
    while True:
        try:
            open_syms = list(account.positions.keys())
            if open_syms:
                changed = False
                all_events: list[dict] = []
                for sym in open_syms:
                    price = await source.fetch_price(session, sym)
                    if price is None:
                        continue
                    for ev in account.update_price(sym, price, price):
                        changed = True
                        all_events.append(ev)
                        print(f"  [SCALP EXIT] {sym:14s} {ev['reason']:4s}  PnL: ${ev['pnl']:>+.2f}")
                if changed:
                    account.save(PAPER_FILE)
                    export_trades_csv(account)
                    append_trades_jsonl(all_events, account)
                    append_equity_snapshot(account)
        except Exception as e:
            log.warning("position_loop error: %s", e)

        await asyncio.sleep(SCALP_CHECK_INTERVAL_SECS)


# ── Loop 2: periodic full re-scan for new signals ────────────

async def scan_loop(
    account: PaperAccount,
    cooldown: dict[str, float],
    session: aiohttp.ClientSession,
) -> None:
    while True:
        try:
            now_dt  = datetime.now(TW_TZ)
            now_str = now_dt.strftime("%Y-%m-%d %H:%M:%S TWN")
            print(f"\n=== SCALP SCAN | {SCALP_INTERVAL} | {now_str} ===")

            now_ts = time.time()
            expired = [s for s, t in cooldown.items() if now_ts - t >= SCALP_COOLDOWN_SECS]
            for s in expired:
                del cooldown[s]

            # ── Gate 2: BTC 趨勢過濾 ─────────────────────────────
            btc_bias, btc_extreme = await fetch_btc_trend(session)
            if btc_extreme:
                print(f"  [SKIP] BTC ADX 極端波動，暫停本輪開倉")
                _now = time.time()
                _next_5m = (int(_now / 300) + 1) * 300 + 5
                await asyncio.sleep(max(_next_5m - time.time(), 10))
                continue
            print(f"  BTC bias: {btc_bias}")

            signals, total = await run_scan("scalp")
            print(f"Scanned: {total} | Signals: {len(signals)}")

            now_ts = time.time()  # refresh after scan completes
            opened  = 0
            skipped = {"blacklist": 0, "score": 0, "rr": 0, "drift": 0,
                       "max_pos": 0, "cooldown": 0, "duplicate": 0, "btc_bias": 0,
                       "funding": 0}

            # Gate 2-8 靜態過濾，收集通過的候選
            candidates = []
            for sig in signals:
                sym = sig["symbol"]
                if len(account.positions) >= SCALP_MAX_POSITIONS:
                    skipped["max_pos"] += 1
                    continue
                if sym in SCALP_SYMBOL_BLOCKLIST:
                    skipped["blacklist"] += 1
                    continue
                if sym in account.positions:
                    skipped["duplicate"] += 1
                    continue
                if now_ts - cooldown.get(sym, 0) < SCALP_COOLDOWN_SECS:
                    skipped["cooldown"] += 1
                    continue
                if sig["score"] < SCALP_MIN_SCORE:
                    skipped["score"] += 1
                    continue
                direction = sig.get("direction", "LONG")
                # 資費極端門檻：多頭擁擠時禁止 LONG，空頭擁擠時禁止 SHORT
                if direction == "LONG"  and sig.get("funding", 0) >  FUNDING_EXTREME:
                    skipped["funding"] += 1
                    continue
                if direction == "SHORT" and sig.get("funding", 0) < -FUNDING_EXTREME:
                    skipped["funding"] += 1
                    continue
                if btc_bias == "BEARISH" and direction == "LONG":
                    skipped["btc_bias"] += 1
                    continue
                if btc_bias == "BULLISH" and direction == "SHORT":
                    skipped["btc_bias"] += 1
                    continue
                if sig.get("risk_reward", 0) < SCALP_MIN_RR:
                    skipped["rr"] += 1
                    continue
                candidates.append(sig)

            # Gate 9: 批次查詢現價，過濾漂移（並行，非串行）
            if candidates:
                live_prices = await asyncio.gather(
                    *[_bingx.fetch_price(session, s["symbol"]) for s in candidates],
                    return_exceptions=True,
                )
            else:
                live_prices = []

            for sig, live_price in zip(candidates, live_prices):
                sym = sig["symbol"]
                if isinstance(live_price, float) and live_price > 0:
                    drift = abs(live_price - sig["entry"]) / sig["entry"]
                    if drift > SCALP_MAX_ENTRY_DRIFT:
                        skipped["drift"] += 1
                        continue

                if account.open_position(sig):
                    cooldown[sym] = now_ts
                    opened += 1
                    strategy = sig.get("strategy", "MA_BREAKOUT")
                    print(
                        f"  [SCALP OPEN] {sym:14s} {strategy:12s}  Score {sig['score']}  "
                        f"R:R {sig.get('risk_reward', 0):.1f}  "
                        f"entry={sig['entry']:.6g}  SL={sig['stop_loss']:.6g}  TP1={sig['target_1']:.6g}"
                    )

            if any(v > 0 for v in skipped.values()):
                parts = [f"{k}={v}" for k, v in skipped.items() if v > 0]
                print(f"  [GATES] skipped: {', '.join(parts)}")

            account.save(PAPER_FILE)
            save_state(account, signals, total)
            export_trades_csv(account)
            append_signals_jsonl(signals, total)
            append_equity_snapshot(account)

            stats = account.get_stats()
            print(
                f"Opened: {opened} | Open positions: {stats['open_positions']} | "
                f"Balance: ${stats['current_balance']:,.2f} ({stats['return_pct']:+.2f}%)"
            )
        except Exception as e:
            log.warning("scan_loop error: %s", e)

        # Sleep until next 5-minute candle close (+5s buffer for exchange to finalise)
        _now = time.time()
        _next_5m = (int(_now / 300) + 1) * 300 + 5
        await asyncio.sleep(max(_next_5m - time.time(), 10))


async def _close_all_positions(account: PaperAccount) -> None:
    """Fetch current prices and force-close every open position at market on shutdown."""
    if not account.positions:
        return
    print(f"\n[SHUTDOWN] 強制平倉 {len(account.positions)} 個持倉...")
    all_events: list[dict] = []
    async with aiohttp.ClientSession() as session:
        for sym in list(account.positions.keys()):
            try:
                price = await source.fetch_price(session, sym)
                if price is None:
                    log.warning("Cannot fetch price for %s — skipping forced close", sym)
                    continue
                ev = account._close(sym, price, "SHUTDOWN")
                all_events.append(ev)
                print(f"  [SHUTDOWN CLOSE] {sym:14s}  PnL: ${ev['pnl']:>+.2f}")
            except Exception as e:
                log.warning("Forced close failed for %s: %s", sym, e)
    if all_events:
        account.save(PAPER_FILE)
        export_trades_csv(account)
        append_trades_jsonl(all_events, account)
        append_equity_snapshot(account)


async def _send_shutdown_report(account: PaperAccount, start_ts: float, stop_ts: float) -> None:
    """Close all positions then send today's trading summary to Discord."""
    await _close_all_positions(account)
    if not SCALP_DISCORD_WEBHOOK:
        return
    try:
        today = datetime.now(TW_TZ).date()
        history_today = [
            t for t in account.history
            if datetime.fromtimestamp(t["close_time"], tz=TW_TZ).date() == today
        ]
        await send_daily_report(
            account.get_stats(), history_today, "bingx", SCALP_DISCORD_WEBHOOK,
            start_time_str=_fmt_tw(start_ts),
            stop_time_str=_fmt_tw(stop_ts),
            duration_str=_duration_str(start_ts, stop_ts),
        )
        print("Discord 關閉報告已發送")
    except Exception as e:
        log.warning("Failed to send shutdown report: %s", e)


async def main_async() -> None:
    global BOT_START_TIME, BOT_STOP_TIME
    BOT_START_TIME = time.time()

    # 每次重啟都重置資金，並清空本輪的 JSONL（archive 已在上次關機時存好）
    account = PaperAccount(SCALP_PAPER_INITIAL_BALANCE, SCALP_PAPER_RISK_PCT)
    account.save(PAPER_FILE)
    for f in (TRADES_JSONL, SIGNALS_JSONL, EQUITY_JSONL, TRADE_CSV):
        if f.exists():
            f.unlink()

    cooldown: dict[str, float] = {}

    # 記錄開機 session
    append_session_log("START", BOT_START_TIME, {"balance": SCALP_PAPER_INITIAL_BALANCE})

    try:
        async with aiohttp.ClientSession() as session:
            await asyncio.gather(
                position_loop(account, session),
                scan_loop(account, cooldown, session),
            )
    except asyncio.CancelledError:
        pass
    finally:
        BOT_STOP_TIME = time.time()
        account.save(PAPER_FILE)

        # 記錄關機 session
        stats    = account.get_stats()
        duration = _duration_str(BOT_START_TIME, BOT_STOP_TIME)
        append_session_log("STOP", BOT_STOP_TIME, {
            "duration": duration,
            "pnl":      stats.get("total_pnl", 0),
            "win_rate": stats.get("win_rate", 0),
            "balance":  stats.get("current_balance", 0),
        })

        # 更新 state 檔案加入 stop_time
        if STATE_FILE.exists():
            try:
                state = json.loads(STATE_FILE.read_text(encoding="utf-8"))
                state["stop_time"] = BOT_STOP_TIME
                STATE_FILE.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")
            except Exception:
                pass

        # 歸檔本次 session 資料
        try:
            archive_session(BOT_START_TIME, BOT_STOP_TIME, stats)
        except Exception as e:
            log.warning("Archive failed: %s", e)

        await _send_shutdown_report(account, BOT_START_TIME, BOT_STOP_TIME)


def main() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
        datefmt="%H:%M:%S",
    )
    for lib in ("aiohttp", "asyncio"):
        logging.getLogger(lib).setLevel(logging.WARNING)

    print("=" * 55)
    print("  Crypto Screener — High-Frequency Scalp Bot (BingX)")
    print(f"  Interval: {SCALP_INTERVAL} | Scan at 5m candle close | "
          f"Position check every {SCALP_CHECK_INTERVAL_SECS}s")
    print("=" * 55)

    try:
        asyncio.run(main_async())
    except KeyboardInterrupt:
        print("\nStopped by user.")


if __name__ == "__main__":
    main()
